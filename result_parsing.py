import os
import re
import openpyxl

aggr_dict = {1:"sum", 2:"average", 3:"max", 4:"wsum", 5:"esum", 6:"gat"}
vec_dict = {128:0, 256:1, 512:2}
vec_dictrev = [128, 256, 512]
# nei_dict = {3:0, 5:1, 10:2, 30:3, 50:4, 100:5}
# nei_dictrev = [3, 5, 10, 30, 50, 100]
nei_dict = {10:0, 20:1, 200:2}
nei_dictrev = [10, 20, 200]
p = '[\d]+[.,\d]+|[\d]*[.][\d]+|[\d]+'

def parsing_result(filepath, filename):
    fin = open(os.path.join(filepath, filename), "rt")
    lines = fin.readlines()
    res = [0 for i in range(3)]
    idx = 0
    for line in lines:
        if "Total time taken" in line:
           idx = 0
        elif "Number of rounds" in line:
           idx = 1
        elif "Total comm (sent+received)" in line:
           idx = 2
        else:
            continue
        if re.search(p, line) is not None:
            catch = re.findall(p, line)
            if len(catch) > 1:
                print("Unexpected string: " + line + " in " + filename)
                exit(0)
            res[idx] = float(catch[0])
    return res

finpath = "/home/taehyun/OpenCheetah/outputs"
fouts = "/home/taehyun/OpenCheetah/results-server.xlsx"
foutc = "/home/taehyun/OpenCheetah/results-client.xlsx"
file_list = os.listdir(finpath)

wbs = openpyxl.Workbook()
wbs.active.title = "sum"
ws1 = wbs["sum"]
ws2 = wbs.create_sheet("average")
ws3 = wbs.create_sheet("max")
ws4 = wbs.create_sheet("wsum")
ws5 = wbs.create_sheet("esum")
ws6 = wbs.create_sheet("gat")
wbc = openpyxl.Workbook()
wbc.active.title = "sum"
wc1 = wbc["sum"]
wc2 = wbc.create_sheet("average")
wc3 = wbc.create_sheet("max")
wc4 = wbc.create_sheet("wsum")
wc5 = wbc.create_sheet("esum")
wc6 = wbc.create_sheet("gat")
for w in [ws1, ws2, ws3, ws4, ws5, ws6, wc1, wc2, wc3, wc4, wc5, wc6]:
    w.merge_cells('A1:E1')
    w.cell(column=1, row=1).value = "Model"
    w.cell(column=1, row=2).value = "layers"
    w.cell(column=2, row=2).value = "fc"
    w.cell(column=3, row=2).value = "vector"
    w.cell(column=4, row=2).value = "nei"
    model_idx = 3
    for ii in range(1, 5): # layers
        for iii in range(1, 3): # fc
            for iiii in range(3): # vector
                for iiiii in range(len(nei_dictrev)): # nei
                    w.cell(column=1, row=model_idx).value = ii
                    w.cell(column=2, row=model_idx).value = iii
                    w.cell(column=3, row=model_idx).value = vec_dictrev[iiii]
                    w.cell(column=4, row=model_idx).value = nei_dictrev[iiiii]
                    model_idx = model_idx + 1
    colidx = 6
    for env in ["One machine", "domestic", "wan", "mobile"]:
        for measurement in ["time(us)", "comm.rounds", "comm.traffic"]:
            w.cell(column=colidx, row=1).value = env + "-" + measurement
            w.cell(column=colidx, row=2).value = "h0"
            w.cell(column=colidx + 1, row=2).value = "aggr_1"
            w.cell(column=colidx + 2, row=2).value = "aggr_2"
            w.cell(column=colidx + 3, row=2).value = "remaining"
            w.cell(column=colidx + 4, row=2).value = "whole"
            w.cell(column=colidx + 5, row=2).value = "Ours"
            colidx = colidx + 6

for filename in file_list:
    w = None
    wb = None
    if filename[4] == "B":
        if re.search(p, filename) is not None:
            catch = re.findall(p, filename)
            fc = int(catch[2])
            vec = int(catch[3])
        res = parsing_result(finpath, filename)
        if 'server' in filename:
            wb = wbs
        elif 'client' in filename:
            wb = wbc
        else:
            continue
        for aggr in aggr_dict.values():
            w = wb[aggr]
            for layer in [1, 2, 3, 4]:
                for nei in nei_dictrev:
                    model_idx = 3 + (layer-1)*len(nei_dictrev)*3*2 + (fc-1)*len(nei_dictrev)*3 + vec_dict[vec]*len(nei_dictrev) + nei_dict[nei]
                    if "wan" in filename:
                        colidx = 6 + 18*2
                    elif "mobile" in filename:
                        colidx = 6 + 18*3
                    elif "domestic" in filename:
                        colidx = 6 + 18
                    else:
                        colidx = 6
                    w.cell(column=colidx, row=model_idx).value = res[0]
                    w.cell(column=colidx+6, row=model_idx).value = res[1]
                    w.cell(column=colidx+12, row=model_idx).value = res[2]

fin = open(os.path.join(finpath, "output.txt"), "rt")
lines = fin.readlines()
for line in lines:
    w = None
    wb = None
    if line[4] != "E":
        continue
    for wb in [wbs, wbc]:
        if re.search(p, line) is not None:
            catch = re.findall(p, line)
            aggr = int(catch[0])
            layer = int(catch[1])
            fc = int(catch[2])
            vec = int(catch[3])
            nei = int(catch[4])
        else:
            continue
        if aggr == 1:
            w = wb["sum"]
        elif aggr == 2:
            w = wb["average"]
        elif aggr == 3:
            w = wb["max"]
        elif aggr == 4:
            w = wb["wsum"]
        elif aggr == 5:
            w = wb["esum"]
        elif aggr == 6:
            w = wb["gat"]
        else:
            continue
        model_idx = 3 + (layer-1)*len(nei_dictrev)*3*2 + (fc-1)*len(nei_dictrev)*3 + vec_dict[vec]*len(nei_dictrev) + nei_dict[nei]
        colidx = 6
        w.cell(column=colidx+3, row=model_idx).value = float(catch[-1]) * 1000000
        w.cell(column=colidx+18*1+3, row=model_idx).value = float(catch[-1]) * 1000000
        w.cell(column=colidx+18*2+3, row=model_idx).value = float(catch[-1]) * 1000000
        w.cell(column=colidx+18*3+3, row=model_idx).value = float(catch[-1]) * 1000000
        w.cell(column=colidx+3+6, row=model_idx).value = 0
        w.cell(column=colidx+18*1+3+6, row=model_idx).value = 0
        w.cell(column=colidx+18*2+3+6, row=model_idx).value = 0
        w.cell(column=colidx+18*3+3+6, row=model_idx).value = 0
        w.cell(column=colidx+3+12, row=model_idx).value = 0
        w.cell(column=colidx+18*1+3+12, row=model_idx).value = 0
        w.cell(column=colidx+18*2+3+12, row=model_idx).value = 0
        w.cell(column=colidx+18*3+3+12, row=model_idx).value = 0

for filename in file_list:
    w = None
    wb = None
    if filename[4] == "E":
        if re.search(p, filename) is not None:
            catch = re.findall(p, filename)
            aggr = int(catch[0])
            layer = int(catch[1])
            fc = int(catch[2])
            vec = int(catch[3])
            nei = int(catch[4])
        if 'server' in filename:
            wb = wbs
        elif 'client' in filename:
            wb = wbc
        else:
            continue
        if aggr != 3:
            continue
        res = parsing_result(finpath, filename)
        print(filename)
        w = wb["max"]
        model_idx = 3 + (layer-1)*len(nei_dictrev)*3*2 + (fc-1)*len(nei_dictrev)*3 + vec_dict[vec]*len(nei_dictrev) + nei_dict[nei]
        if "wan" in filename:
            colidx = 6 + 18*2
        elif "mobile" in filename:
            colidx = 6 + 18*3
        elif "domestic" in filename:
            colidx = 6 + 18
        else:
            colidx = 6
        w.cell(column=colidx+3, row=model_idx).value = res[0]
        w.cell(column=colidx+3+6, row=model_idx).value = res[1]
        w.cell(column=colidx+3+12, row=model_idx).value = res[2]

for filename in file_list:
    w = None
    wb = None
    if filename[4] == "B":
        continue
    if 'server' in filename:
        wb = wbs
    elif 'client' in filename:
        wb = wbc
    else:
        continue
    if re.search(p, filename) is not None:
        catch = re.findall(p, filename)
        aggr = int(catch[0])
        layer = int(catch[1])
        fc = int(catch[2])
        vec = int(catch[3])
        nei = int(catch[4])
    else:
        continue
    if aggr == 1:
        w = wb["sum"]
    elif aggr == 2:
        w = wb["average"]
    elif aggr == 3:
        w = wb["max"]
    elif aggr == 4:
        w = wb["wsum"]
    elif aggr == 5:
        w = wb["esum"]
    elif aggr == 6:
        w = wb["gat"]
    else:
        continue
    if "wan" in filename:
        colidx = 6 + 18*2
    elif "mobile" in filename:
        colidx = 6 + 18*3
    elif "domestic" in filename:
        colidx = 6 + 18
    else:
        colidx = 6
    if filename[4] == "A":
        colidx = colidx + 4
    elif filename[4] == "C":
        colidx = colidx + 1
    elif filename[4] == "D":
        colidx = colidx + 2
    else:
        continue

    model_idx = 3 + (layer-1)*len(nei_dictrev)*3*2 + (fc-1)*len(nei_dictrev)*3 + vec_dict[vec]*len(nei_dictrev) + nei_dict[nei]
    res = parsing_result(finpath, filename)
    w.cell(column=colidx, row=model_idx).value = res[0]
    w.cell(column=colidx+6, row=model_idx).value = res[1]
    w.cell(column=colidx+12, row=model_idx).value = res[2]
    if filename[4] == "C" or filename[4] == "D":
        model_idx = model_idx + 18
        w.cell(column=colidx, row=model_idx).value = res[0]
        w.cell(column=colidx+6, row=model_idx).value = res[1]
        w.cell(column=colidx+12, row=model_idx).value = res[2]

wbs.save(fouts)
wbc.save(foutc)